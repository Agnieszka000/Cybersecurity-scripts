# Script that takes a bulk of IP addresses from an Excel file (.xlsx) and checks on VirusTotal if they're malicious, using API key.
# To perform a scan your API key is needed.
# This script was generated mostly by Claud AI.


import requests
import openpyxl
import re
from time import sleep
from tqdm import tqdm   # Library to show the status bar.

def extract_ips_from_excel(file_path):
    print("Extracting IP addresses from Excel file...")
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
    ip_addresses = set()
    for row in tqdm(sheet.iter_rows(), total=sheet.max_row, desc="Processing rows"):
        for cell in row:
            if cell.value:
                found_ips = re.findall(ip_pattern, str(cell.value))
                ip_addresses.update(found_ips)
    print(f"Found {len(ip_addresses)} unique IP addresses.")
    return list(ip_addresses)

def check_ip_virustotal(ip, api_key):
    url = f'https://www.virustotal.com/api/v3/ip_addresses/{ip}'
    # VirusTotal uses 'x-apikey' in headers instead of query params
    headers = {
        'Accept': 'application/json',
        'x-apikey': api_key
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        # Parsing VirusTotal's response structure
        result = response.json()['data']['attributes']
        return {
            'ip': ip,
            # Extracting relevant information from VirusTotal response
            'malicious': result['last_analysis_stats']['malicious'],
            'suspicious': result['last_analysis_stats']['suspicious'],
            'harmless': result['last_analysis_stats']['harmless'],
            'undetected': result['last_analysis_stats']['undetected'],
            'country': result.get('country', 'Unknown')
        }
    else:
        return {'ip': ip, 'error': f'API request failed with status code {response.status_code}'}

def bulk_check_ips(ips, api_key):
    results = []
    # Updated message to reflect use of VirusTotal
    print("Checking IP addresses with VirusTotal...")
    for ip in tqdm(ips, desc="Checking IPs"):
        result = check_ip_virustotal(ip, api_key)
        results.append(result)
        # Sleep time of 15 seconds due to VirusTotal's rate limits
        sleep(15)  # VirusTotal has a rate limit of 4 requests per minute for free accounts
    return results

# Usage
# Path to the Excel .xlsx file with IP addresses: 
excel_file_path = '<EXCEL_FILE_PATH_HERE>'
# VirusTotal API key
api_key = '<YOUR_API_KEY_HERE>'  # Replace with your actual API key

ips = extract_ips_from_excel(excel_file_path)
results = bulk_check_ips(ips, api_key)

print("\nResults:")
for result in results:
    if 'error' in result:
        print(f"{result['ip']}: Error - {result['error']}")
    else:
        print(f"{result['ip']}: Malicious: {result['malicious']}, "
              f"Suspicious: {result['suspicious']}, Harmless: {result['harmless']}, "
              f"Undetected: {result['undetected']}, Country: {result['country']}")

print("IP checking complete.")